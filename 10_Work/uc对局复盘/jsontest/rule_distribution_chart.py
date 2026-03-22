#!/usr/bin/env python3
"""
基于 label_id 精确分析【有待提升】规则命中分布，并生成 Plotly 可视化图表。
"""
import json
from collections import Counter, defaultdict
from openpyxl import load_workbook
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX_PATH = "result.xlsx"

# label_id → 规则编号 & 描述
RULE_LABEL_MAP = {
    14: ("规则1", "防反触发/防御<0.2"),
    17: ("规则2", "变招先手/防反>3"),
    18: ("规则3", "脱出闪反打=0"),
    21: ("规则4", "倒地受击>2×均值"),
    25: ("规则5", "体力消耗>2×均值"),
    41: ("规则6", "防御暂停恢复<对手"),
    43: ("规则7", "炁满伤害<对手"),
    44: ("规则8", "身外身冷却<对手"),
    42: ("规则9", "无法脱出受击<对手"),
}

ALL_RULES_ORDERED = [14, 17, 18, 21, 25, 41, 43, 44, 42]


def load_data():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, filename, response_text = row[0], row[1], row[2]
        if not response_text:
            continue
        try:
            data = json.loads(response_text)
        except json.JSONDecodeError:
            continue

        for sec in data.get("sections", []):
            content = sec.get("content", "")
            if "有待提升" not in content:
                continue

            label_id = sec.get("label_id", -1)
            is_rule = sec.get("is_rule", False)
            brief = content.replace("【有待提升】", "").replace('\n', ' ').strip()

            # 分类
            if label_id in RULE_LABEL_MAP:
                rule_num, desc = RULE_LABEL_MAP[label_id]
                category = f"{rule_num}: {desc}"
                cat_type = "规则命中"
            elif label_id == -1 or label_id == 0:
                category = "非规则: 绝技使用建议"
                cat_type = "非规则路径"
            else:
                category = f"非规则: improveSon(id={label_id})"
                cat_type = "非规则路径"

            records.append({
                "filename": filename,
                "label_id": label_id,
                "is_rule": is_rule,
                "category": category,
                "cat_type": cat_type,
                "brief": brief[:80],
            })
            break

    return records


def main():
    records = load_data()
    total = len(records)

    # ── 统计 ──
    cat_counter = Counter(r["category"] for r in records)

    # 按规则顺序排列
    ordered_cats = []
    ordered_counts = []
    for lid in ALL_RULES_ORDERED:
        rule_num, desc = RULE_LABEL_MAP[lid]
        cat = f"{rule_num}: {desc}"
        ordered_cats.append(cat)
        ordered_counts.append(cat_counter.get(cat, 0))

    # 非规则路径
    non_rule_cats = sorted([c for c in cat_counter if c.startswith("非规则")])
    for cat in non_rule_cats:
        ordered_cats.append(cat)
        ordered_counts.append(cat_counter[cat])

    # ── 颜色 ──
    colors = []
    for cat in ordered_cats:
        if cat.startswith("规则1"):
            colors.append("#FF6B6B")
        elif cat.startswith("规则2"):
            colors.append("#FFA07A")
        elif cat.startswith("规则3"):
            colors.append("#4ECDC4")
        elif cat.startswith("规则4"):
            colors.append("#45B7D1")
        elif cat.startswith("规则5"):
            colors.append("#96CEB4")
        elif cat.startswith("规则6"):
            colors.append("#FFEAA7")
        elif cat.startswith("规则7"):
            colors.append("#DDA0DD")
        elif cat.startswith("规则8"):
            colors.append("#98D8C8")
        elif cat.startswith("规则9"):
            colors.append("#F7DC6F")
        elif "绝技" in cat:
            colors.append("#A29BFE")
        else:
            colors.append("#B8B8B8")

    # ── 创建图表 ──
    fig = make_subplots(
        rows=1, cols=2,
        specs=[[{"type": "bar"}, {"type": "pie"}]],
        subplot_titles=[
            "各规则命中数量",
            "规则命中占比"
        ],
        column_widths=[0.6, 0.4],
    )

    # 柱状图
    fig.add_trace(
        go.Bar(
            y=ordered_cats,
            x=ordered_counts,
            orientation='h',
            marker_color=colors,
            text=[f"{c} ({c/total*100:.1f}%)" if c > 0 else "0" for c in ordered_counts],
            textposition='outside',
            textfont=dict(size=12),
        ),
        row=1, col=1,
    )

    # 饼图 - 只显示有数据的
    pie_labels = []
    pie_values = []
    pie_colors = []
    for cat, cnt, clr in zip(ordered_cats, ordered_counts, colors):
        if cnt > 0:
            pie_labels.append(cat)
            pie_values.append(cnt)
            pie_colors.append(clr)

    fig.add_trace(
        go.Pie(
            labels=pie_labels,
            values=pie_values,
            marker=dict(colors=pie_colors),
            textinfo="label+percent",
            textfont=dict(size=11),
            hole=0.3,
        ),
        row=1, col=2,
    )

    fig.update_layout(
        title=dict(
            text=f"📊 【有待提升】规则命中分布 (共 {total} 条)",
            font=dict(size=18),
        ),
        height=650,
        width=1400,
        showlegend=False,
        margin=dict(l=250, r=50, t=80, b=50),
    )

    fig.update_xaxes(title_text="命中数量", row=1, col=1)
    fig.update_yaxes(autorange="reversed", row=1, col=1)

    output_path = "rule_distribution.html"
    fig.write_html(output_path)
    print(f"✅ 图表已保存到 {output_path}")

    # ── 文本汇总 ──
    print(f"\n{'='*70}")
    print(f"📊 【有待提升】规则命中分布 (共 {total} 条)")
    print(f"{'='*70}")
    print(f"{'类别':<35s} {'数量':>4s} {'占比':>7s}")
    print(f"{'─'*55}")
    for cat, cnt in zip(ordered_cats, ordered_counts):
        pct = cnt / total * 100
        bar = "█" * int(pct / 2)
        print(f"  {cat:<33s} {cnt:>4d}  {pct:>5.1f}%  {bar}")
    print(f"{'─'*55}")
    print(f"  {'合计':<33s} {total:>4d}  100.0%")

    # ── 关键发现 ──
    print(f"\n{'='*70}")
    print("🔍 关键发现")
    print(f"{'='*70}")

    rule_hit = sum(1 for r in records if r["cat_type"] == "规则命中")
    non_rule_hit = total - rule_hit
    print(f"1. 规则1-9 命中: {rule_hit}/{total} ({rule_hit/total*100:.1f}%)")
    print(f"   非规则路径:   {non_rule_hit}/{total} ({non_rule_hit/total*100:.1f}%)")
    print()

    # 命中的规则
    hit_rules = [(lid, cat_counter.get(f"{RULE_LABEL_MAP[lid][0]}: {RULE_LABEL_MAP[lid][1]}", 0))
                 for lid in ALL_RULES_ORDERED]
    hit_rules_nz = [(lid, cnt) for lid, cnt in hit_rules if cnt > 0]
    miss_rules = [lid for lid, cnt in hit_rules if cnt == 0]

    print(f"2. 有命中的规则 ({len(hit_rules_nz)} 条):")
    for lid, cnt in hit_rules_nz:
        r, d = RULE_LABEL_MAP[lid]
        print(f"   {r}: {d} → {cnt} 条 ({cnt/total*100:.1f}%)")

    print(f"\n3. 完全没命中的规则 ({len(miss_rules)} 条):")
    for lid in miss_rules:
        r, d = RULE_LABEL_MAP[lid]
        print(f"   {r}: {d} → 0 条")

    top2 = sorted(hit_rules_nz, key=lambda x: -x[1])[:2]
    top2_total = sum(c for _, c in top2)
    print(f"\n4. Top2 规则占比: {top2_total}/{total} = {top2_total/total*100:.1f}%")
    for lid, cnt in top2:
        r, d = RULE_LABEL_MAP[lid]
        print(f"   {r} ({d}): {cnt} 条")


if __name__ == "__main__":
    main()
