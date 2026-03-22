#!/usr/bin/env python3
"""
分析 result.xlsx 中每条【有待提升】命中了哪条规则。

代码逻辑（来自 analysis.go）：
  规则1-9 是 if-else if 链，命中后 is_rule=True，输出 special_texts 模板文字。
  都不命中 → 走 improveSon 逻辑，用 {player_val}/{opponent_val} 替换，输出带数字的文本。
  特殊：myImproveWV >= opImproveWV → betterOpp 文本（"全面优于对手"）

通过返回文本的关键词 + is_rule 字段来反推命中了哪个路径。
"""
import json
import re
from collections import Counter, defaultdict
from openpyxl import load_workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX_PATH = "result.xlsx"

# ── 规则关键词映射 ──
# 每个规则对应 special_texts 的 key，我们通过文本特征来区分
RULE_PATTERNS = [
    # (规则名, 关键词列表 - 任一匹配即可)
    ("规则1: 防反触发/防御<0.2 (extremeDefAtk)",
     ["防反", "防御反击"]),
    ("规则2: 变招先手/防反>3 (extremeSwitchLead)",
     ["变招", "先手"]),
    ("规则3: 脱出闪反打=0 (extremeEscape)",
     ["脱出", "闪反打", "闪反"]),
    ("规则4: 倒地受击>2x均值 (extremePressure)",
     ["倒地", "受击", "起身"]),
    ("规则5: 体力消耗>2x均值 (extremeNoRecovery)",
     ["体力消耗", "体力恢复", "体力管理"]),
    ("规则6: 防御暂停体力恢复<对手 (extremeDefRecovery)",
     ["防御暂停", "暂停体力"]),
    ("规则7: 炁满伤害<对手 (extremeFullQiDmg)",
     ["炁满", "炁", "气满"]),
    ("规则8: 身外身冷却<对手 (extremeCloneCooldown)",
     ["身外身", "冷却"]),
    ("规则9: 无法脱出受击<对手 (extremeNoEscapeHit)",
     ["无法脱出", "不可脱出"]),
]


def classify_rule(text: str, is_rule: bool) -> str:
    """根据文本内容和 is_rule 标志判断命中了哪条规则。"""
    # 去掉标题
    clean = text.replace("【有待提升】", "").strip()

    if not is_rule:
        return "非规则路径 (improveSon 带数字)"

    # 检查是否是 betterOpp（全面优于对手的情况）
    if any(kw in clean for kw in ["全面优于", "整体优于", "表现更好", "优于对手"]):
        return "betterOpp: 全面优于对手"

    # 尝试匹配规则 1-9
    for rule_name, keywords in RULE_PATTERNS:
        if any(kw in clean for kw in keywords):
            return rule_name

    # 兜底：有数字但走了规则路径？
    if re.search(r'\d', clean):
        return "非规则路径 (improveSon 带数字)"

    return f"未分类 (is_rule={is_rule})"


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    rule_counter = Counter()
    rule_details = defaultdict(list)  # rule -> [(filename, text_brief)]
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, filename, response_text = row[0], row[1], row[2]
        if not response_text:
            continue

        try:
            data = json.loads(response_text)
        except json.JSONDecodeError:
            continue

        sections = data.get("sections", [])

        weakness_text = ""
        is_rule = False
        for sec in sections:
            content = sec.get("content", "")
            if "有待提升" in content:
                weakness_text = content
                is_rule = sec.get("is_rule", False)
                break

        if not weakness_text:
            continue

        total += 1
        rule = classify_rule(weakness_text, is_rule)
        rule_counter[rule] += 1

        brief = weakness_text.replace('\n', ' ').strip()
        if len(brief) > 120:
            brief = brief[:120] + "..."
        rule_details[rule].append((filename, brief))

    # ── 输出汇总 ──
    print(f"{'='*90}")
    print(f"📊 【有待提升】规则命中分布  (共 {total} 条)")
    print(f"{'='*90}")
    print()
    print(f"{'规则':<50s} {'数量':>4s} {'占比':>7s}")
    print(f"{'─'*65}")

    for rule, count in rule_counter.most_common():
        pct = count / total * 100
        bar = "█" * int(pct / 2)
        print(f"  {rule:<48s} {count:>4d}  {pct:>5.1f}%  {bar}")

    print(f"{'─'*65}")
    print(f"  {'合计':<48s} {total:>4d}  100.0%")
    print()

    # ── 每个规则的详细列表 ──
    print(f"{'='*90}")
    print(f"📋 各规则命中明细")
    print(f"{'='*90}")

    for rule, count in rule_counter.most_common():
        print(f"\n▶ {rule}  [{count}条]")
        print(f"  {'─'*80}")
        for fname, brief in rule_details[rule]:
            print(f"  [{fname}] {brief}")


if __name__ == "__main__":
    main()
