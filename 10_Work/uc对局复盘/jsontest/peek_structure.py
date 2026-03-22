#!/usr/bin/env python3
"""探查 result.xlsx 中返回数据的结构"""
import json
from openpyxl import load_workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = load_workbook("result.xlsx")
ws = wb.active

# 只看第一条
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    seq, filename, response_text = row[0], row[1], row[2]
    data = json.loads(response_text)
    
    print(f"顶层键: {list(data.keys())}")
    sections = data.get("sections", [])
    print(f"sections 数量: {len(sections)}")
    
    for i, sec in enumerate(sections):
        content = sec.get("content", "")
        is_rule = sec.get("is_rule", "")
        print(f"\n--- Section {i} (is_rule={is_rule}) ---")
        print(content[:500])
        print("...")
