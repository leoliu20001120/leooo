#!/usr/bin/env python3
"""
逐个上传 JSON 文件到服务器，并将返回结果写入 Excel 文件。
"""

import os
import json
import glob
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ── 配置 ──────────────────────────────────────────────
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
URL = "http://30.189.253.210:8080/fight/report"
OUTPUT_XLSX = os.path.join(WORK_DIR, "result.xlsx")
TIMEOUT = 60  # 每个请求的超时时间（秒）


def collect_json_files(directory: str) -> list[str]:
    """收集目录下所有数字命名的 JSON 文件，并按数字顺序排列。"""
    pattern = os.path.join(directory, "*.json")
    files = glob.glob(pattern)
    # 只保留文件名是 纯数字.json 的
    json_files = []
    for f in files:
        basename = os.path.splitext(os.path.basename(f))[0]
        if basename.isdigit():
            json_files.append(f)
    # 按数字排序
    json_files.sort(key=lambda f: int(os.path.splitext(os.path.basename(f))[0]))
    return json_files


def upload_json(filepath: str) -> str:
    """上传单个 JSON 文件，返回服务器响应文本。"""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    try:
        resp = requests.post(URL, json=data, timeout=TIMEOUT)
        return resp.text
    except requests.RequestException as e:
        return f"[ERROR] {e}"


def write_excel(results: list[tuple[int, str, str]], output_path: str):
    """
    将结果写入 Excel 文件。
    results: [(序号, 文件名, 返回信息), ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "上传结果"

    # ── 表头 ──
    headers = ["序号", "文件名", "返回信息"]
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # ── 数据行 ──
    for row_idx, (seq, filename, response) in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=seq).border = thin_border
        ws.cell(row=row_idx, column=2, value=filename).border = thin_border
        ws.cell(row=row_idx, column=3, value=response).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

    # ── 列宽 ──
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 80

    wb.save(output_path)
    print(f"\n✅ 结果已写入: {output_path}")


def main():
    json_files = collect_json_files(WORK_DIR)
    if not json_files:
        print("❌ 未找到任何数字命名的 JSON 文件")
        return

    print(f"📂 共找到 {len(json_files)} 个 JSON 文件")
    print(f"🌐 上传地址: {URL}\n")

    results = []
    for idx, filepath in enumerate(json_files, start=1):
        filename = os.path.basename(filepath)
        print(f"[{idx:>3}/{len(json_files)}] 上传 {filename} ...", end=" ", flush=True)
        response = upload_json(filepath)
        # 截断过长的响应用于终端显示
        display = response[:100] + "..." if len(response) > 100 else response
        print(f"→ {display}")
        results.append((idx, filename, response))

    write_excel(results, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
