#!/usr/bin/env python3
"""
分析 result.xlsx 中【有待提升】部分是否包含具体指标数字。
"""
import json
import re
from openpyxl import load_workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX_PATH = "result.xlsx"


def has_specific_numbers(text: str) -> bool:
    """
    判断文本中是否包含具体指标数字。
    匹配：百分比、小数、带单位数字等。
    排除：纯文字描述型内容。
    """
    # 去掉开头的 【有待提升】 标题
    text = re.sub(r'【.*?】', '', text)
    
    # 百分比
    if re.search(r'\d+\.?\d*\s*%', text):
        return True
    # 带小数点的数字（如 0.85, 2.3, 5.5）
    if re.search(r'\d+\.\d+', text):
        return True
    # X分（评分）
    if re.search(r'\d+\.?\d*\s*分', text):
        return True
    # 带单位的数字
    if re.search(r'\d+\.?\d*\s*[秒次局回合]', text):
        return True
    # X:Y 比分
    if re.search(r'\d+\s*:\s*\d+', text):
        return True
    # X/Y 比值
    if re.search(r'\d+\s*/\s*\d+', text):
        return True
    return False


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    total = 0
    no_number_count = 0
    has_number_count = 0
    no_number_details = []
    has_number_details = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, filename, response_text = row[0], row[1], row[2]
        if not response_text:
            continue

        try:
            data = json.loads(response_text)
        except json.JSONDecodeError:
            print(f"⚠️ {filename}: JSON 解析失败")
            continue

        sections = data.get("sections", [])
        
        # 找【有待提升】section
        weakness_text = ""
        for sec in sections:
            content = sec.get("content", "")
            if "有待提升" in content:
                weakness_text = content
                break

        if not weakness_text:
            print(f"⚠️ {filename}: 未找到【有待提升】")
            continue

        total += 1

        if has_specific_numbers(weakness_text):
            has_number_count += 1
            has_number_details.append((filename, weakness_text))
        else:
            no_number_count += 1
            no_number_details.append((filename, weakness_text))

    print(f"{'='*80}")
    print(f"📊 【有待提升】中是否包含具体指标数字 — 分析结果")
    print(f"{'='*80}")
    print(f"总对局数:            {total}")
    print(f"✅ 有具体指标数字:    {has_number_count}  ({has_number_count/total*100:.1f}%)")
    print(f"❌ 没有具体指标数字:  {no_number_count}  ({no_number_count/total*100:.1f}%)")
    print()

    # ── 没有数字的列表 ──
    if no_number_details:
        print(f"{'─'*80}")
        print(f"❌ 没有具体指标数字的对局 ({no_number_count}个):")
        print(f"{'─'*80}")
        for fname, text in no_number_details:
            # 去掉换行，截取展示
            brief = text.replace('\n', ' ').strip()
            if len(brief) > 200:
                brief = brief[:200] + "..."
            print(f"  [{fname}]")
            print(f"    {brief}")
            print()

    # ── 有数字的列表 ──
    if has_number_details:
        print(f"{'─'*80}")
        print(f"✅ 有具体指标数字的对局 ({has_number_count}个):")
        print(f"{'─'*80}")
        for fname, text in has_number_details:
            brief = text.replace('\n', ' ').strip()
            if len(brief) > 200:
                brief = brief[:200] + "..."
            print(f"  [{fname}]")
            print(f"    {brief}")
            print()


if __name__ == "__main__":
    main()
