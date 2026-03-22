#!/usr/bin/env python3
"""
基于 label_id 精确分析【有待提升】命中了哪条规则。

规则 -> label_id 映射（来自 analysis.go）：
  规则1: 防反触发次数/防御成功次数 < 0.2     → behId = rankId_14  → label_id 含 14
  规则2: 变招先手次数/防反触发次数 > 3        → behId = rankId_17  → label_id 含 17
  规则3: 脱出闪反打成功为0                     → behId = rankId_18  → label_id 含 18
  规则4: 倒地受击次数 > 2*段位均值             → behId = rankId_21  → label_id 含 21
  规则5: 体力消耗频数 > 2*段位均值             → behId = rankId_25  → label_id 含 25
  规则6: 防御暂停体力恢复时间 < 对手           → behId = rankId_41  → label_id 含 41
  规则7: 炁满时累积伤害量 < 对手               → behId = rankId_43  → label_id 含 43
  规则8: 身外身冷却空转时间 < 对手             → behId = rankId_44  → label_id 含 44
  规则9: 无法脱出总受击时间 < 对手             → behId = rankId_42  → label_id 含 42

  非规则路径: is_rule=False, label_id 是 improveSon 子指标 id
"""
import json
from collections import Counter, defaultdict
from openpyxl import load_workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX_PATH = "result.xlsx"

# label_id → 规则编号 & 描述
# 注意: label_id 在 Go 代码中是 ReadConfigMustInt("barchart_behId", "rankId_14")
# 实际存的就是一个整数，我们直接用 label_id 的值来匹配
RULE_LABEL_MAP = {
    14: ("规则1", "防反触发/防御 < 0.2", "extremeDefAtk"),
    17: ("规则2", "变招先手/防反 > 3", "extremeSwitchLead"),
    18: ("规则3", "脱出闪反打 = 0", "extremeEscape"),
    21: ("规则4", "倒地受击 > 2×段位均值", "extremePressure"),
    25: ("规则5", "体力消耗 > 2×段位均值", "extremeNoRecovery"),
    41: ("规则6", "防御暂停体力恢复 < 对手", "extremeDefRecovery"),
    43: ("规则7", "炁满累积伤害 < 对手", "extremeFullQiDmg"),
    44: ("规则8", "身外身冷却空转 < 对手", "extremeCloneCooldown"),
    42: ("规则9", "无法脱出受击 < 对手", "extremeNoEscapeHit"),
}

# 所有规则按编号排序，用于展示
ALL_RULES_ORDERED = [14, 17, 18, 21, 25, 41, 43, 44, 42]


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    rule_counter = Counter()
    rule_details = defaultdict(list)
    non_rule_counter = Counter()  # 非规则路径的 label_id
    non_rule_details = defaultdict(list)
    total = 0
    total_rule = 0
    total_non_rule = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, filename, response_text = row[0], row[1], row[2]
        if not response_text:
            continue

        try:
            data = json.loads(response_text)
        except json.JSONDecodeError:
            continue

        sections = data.get("sections", [])

        for sec in sections:
            content = sec.get("content", "")
            if "有待提升" not in content:
                continue

            total += 1
            is_rule = sec.get("is_rule", False)
            label_id = sec.get("label_id", 0)

            brief = content.replace("【有待提升】", "").replace('\n', ' ').strip()
            if len(brief) > 100:
                brief = brief[:100] + "..."

            if is_rule:
                total_rule += 1
                rule_counter[label_id] += 1
                rule_details[label_id].append((filename, brief))
            else:
                total_non_rule += 1
                non_rule_counter[label_id] += 1
                non_rule_details[label_id].append((filename, brief))
            break

    # ── 输出汇总 ──
    print(f"{'='*95}")
    print(f"📊 【有待提升】规则命中分布 (基于 label_id)  共 {total} 条")
    print(f"   其中 is_rule=True: {total_rule} 条,  is_rule=False: {total_non_rule} 条")
    print(f"{'='*95}")
    print()

    # ── 规则路径 (is_rule=True) ──
    print(f"┌─────────────────────────────────────────────────────────────────────────┐")
    print(f"│  📌 规则路径 (is_rule=True)  共 {total_rule} 条                          │")
    print(f"├─────┬──────────────────────────────────┬──────┬────────┬────────────────┤")
    print(f"│  #  │ 规则条件                         │ 数量 │  占比  │ 分布           │")
    print(f"├─────┼──────────────────────────────────┼──────┼────────┼────────────────┤")

    for lid in ALL_RULES_ORDERED:
        rule_num, desc, key = RULE_LABEL_MAP[lid]
        cnt = rule_counter.get(lid, 0)
        pct = cnt / total * 100 if total > 0 else 0
        pct_of_rule = cnt / total_rule * 100 if total_rule > 0 else 0
        bar = "█" * int(pct_of_rule / 3)
        mark = " ← 0" if cnt == 0 else ""
        print(f"│ {rule_num} │ {desc:<32s} │ {cnt:>4d} │ {pct:>5.1f}% │ {bar:<14s} │{mark}")

    print(f"├─────┼──────────────────────────────────┼──────┼────────┼────────────────┤")
    print(f"│ 合计│                                  │ {total_rule:>4d} │ {total_rule/total*100:>5.1f}% │                │")
    print(f"└─────┴──────────────────────────────────┴──────┴────────┴────────────────┘")
    print()

    # 检查是否有意外的 label_id (is_rule=True 但 label_id 不在映射中)
    unexpected = {lid: cnt for lid, cnt in rule_counter.items() if lid not in RULE_LABEL_MAP}
    if unexpected:
        print(f"  ⚠️  is_rule=True 但 label_id 不在预期映射中: {unexpected}")
        print()

    # ── 非规则路径 (is_rule=False) ──
    if total_non_rule > 0:
        print(f"┌─────────────────────────────────────────────────────────────────────────┐")
        print(f"│  📎 非规则路径 (is_rule=False, improveSon)  共 {total_non_rule} 条        │")
        print(f"├────────────┬──────┬──────────────────────────────────────────────────────┤")
        print(f"│ label_id   │ 数量 │ 示例文本                                           │")
        print(f"├────────────┼──────┼──────────────────────────────────────────────────────┤")
        for lid, cnt in non_rule_counter.most_common():
            example = non_rule_details[lid][0][1][:50] + "..." if non_rule_details[lid] else ""
            print(f"│ {lid:>10d} │ {cnt:>4d} │ {example:<50s} │")
        print(f"├────────────┼──────┼──────────────────────────────────────────────────────┤")
        print(f"│ 合计       │ {total_non_rule:>4d} │                                                    │")
        print(f"└────────────┴──────┴──────────────────────────────────────────────────────┘")
    print()

    # ── 各规则命中明细 ──
    print(f"{'='*95}")
    print(f"📋 各规则命中明细")
    print(f"{'='*95}")

    for lid in ALL_RULES_ORDERED:
        rule_num, desc, key = RULE_LABEL_MAP[lid]
        cnt = rule_counter.get(lid, 0)
        if cnt == 0:
            print(f"\n▶ {rule_num} (label_id={lid}): {desc}  [0条]")
            continue
        print(f"\n▶ {rule_num} (label_id={lid}): {desc}  [{cnt}条]")
        print(f"  {'─'*85}")
        for fname, brief in rule_details[lid]:
            print(f"  [{fname}] {brief}")

    if total_non_rule > 0:
        print(f"\n▶ 非规则路径 (improveSon)  [{total_non_rule}条]")
        print(f"  {'─'*85}")
        for lid, details in non_rule_details.items():
            for fname, brief in details:
                print(f"  [{fname}] (label_id={lid}) {brief}")


if __name__ == "__main__":
    main()
