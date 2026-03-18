import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('improvement_templates.xlsx')
ws_t = wb['提升点文案模版']
ws_p = wb['待用']

# 收集数据
templates = defaultdict(list)
for r in ws_t.iter_rows(min_row=2, max_row=ws_t.max_row, values_only=True):
    if r[0] and r[1]:
        templates[r[0]].append(r[1])

suggestions = defaultdict(list)
dim_map = {}
for r in ws_p.iter_rows(min_row=2, max_row=ws_p.max_row, values_only=True):
    if r[1] and r[2]:
        suggestions[r[1]].append(r[2])
        dim_map[r[1]] = r[0]

# 删除旧sheet，新建
if '文案x建议' in wb.sheetnames:
    del wb['文案x建议']
ws = wb.create_sheet('文案x建议')
ws.cell(row=1, column=1, value='所属维度名称')
ws.cell(row=1, column=2, value='指标名称')
ws.cell(row=1, column=3, value='文案模版')
ws.cell(row=1, column=4, value='建议')

row_idx = 2
matched = sorted(set(templates.keys()) & set(suggestions.keys()))
for name in matched:
    dim = dim_map.get(name, '')
    for tmpl in templates[name]:
        for sug in suggestions[name]:
            ws.cell(row=row_idx, column=1, value=dim)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=tmpl)
            ws.cell(row=row_idx, column=4, value=sug)
            row_idx += 1

wb.save('improvement_templates.xlsx')
print(f'Done! Restored {row_idx - 2} rows with separate template and suggestion columns.')
