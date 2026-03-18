# -*- coding: utf-8 -*-
"""
高光点模块 (Highlights Module)

核心逻辑：
1. 基于 Δ_dimension = 玩家维度分 − 对手维度分，选取优势维度 max(Δ_dimension)
2. 在优势维度下，选取 max(玩家本人加权值) 的子指标作为最佳表现指标
3. 三段式输出：
   ● Part 1: 展示用户数据亮点（模版化输出，根据雷达图维度 + 胜负区分不同模版）
   ● Part 2: 展示段位平均水平（从段位指标知识库查询该指标的段位平均表现）
   ● Part 3: 推送详细数据链接（跳转至本局完整数据面板）

数据来源：
  - 文案模版 & 指标信息 → highlight_templates.xlsx
  - 段位平均水平 → rank_average.xlsx
"""

import os
import random
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field

import openpyxl

from radar_interpretation import (
    RadarDimension,
    RadarSubDimension,
    DIMENSION_NAMES,
    SUB_DIMENSION_NAMES,
    DIMENSION_SUB_DIMENSIONS,
    WEIGHTED_OFFSET,
    RadarChartData,
    DimensionData,
    SubDimensionData,
    DimensionDelta,
    DimensionDeltaResult,
    parse_radar_chart,
    compute_dimension_deltas,
)


# ================================================================
#  Excel 文件路径（相对于本文件所在目录）
# ================================================================

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_TEMPLATES_XLSX = os.path.join(_BASE_DIR, "highlight_templates.xlsx")
_RANK_AVG_XLSX = os.path.join(_BASE_DIR, "rank_average.xlsx")


# ================================================================
#  从 Excel 加载数据
# ================================================================

def _load_indicator_info(xlsx_path: str) -> Tuple[
    Dict[int, str],   # SUB_DIM_DISPLAY_NAMES: {sub_dim_id: display_name}
    Dict[int, str],   # SUB_DIM_UNITS:         {sub_dim_id: unit}
    set,               # TIME_BASED_SUB_DIMS:   {sub_dim_id, ...}
]:
    """
    从 highlight_templates.xlsx 的「指标信息」sheet 加载指标元数据。

    Excel 列: sub_dim_id | 指标名称 | 单位 | 所属维度ID | 所属维度名称 | 是否时间类(ms→秒)
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["指标信息"]

    display_names: Dict[int, str] = {}
    units: Dict[int, str] = {}
    time_based: set = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        sub_dim_id = int(row[0])
        name = str(row[1] or "")
        unit = str(row[2] or "")
        is_time = str(row[5] or "").strip()

        display_names[sub_dim_id] = name
        units[sub_dim_id] = unit
        if is_time == "是":
            time_based.add(sub_dim_id)

    wb.close()
    return display_names, units, time_based


def _load_highlight_templates(xlsx_path: str) -> Dict[int, Dict[str, List[str]]]:
    """
    从 highlight_templates.xlsx 的「亮点文案模版」sheet 加载文案模版。

    Excel 列: sub_dim_id | 指标名称 | 胜负(win/loss) | 文案模版
    返回: { sub_dim_id: { "win": [tpl, ...], "loss": [tpl, ...] } }
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["亮点文案模版"]

    templates: Dict[int, Dict[str, List[str]]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        sub_dim_id = int(row[0])
        result = str(row[2] or "").strip().lower()  # "win" or "loss"
        tpl_text = str(row[3] or "").strip()

        if not tpl_text or result not in ("win", "loss"):
            continue

        if sub_dim_id not in templates:
            templates[sub_dim_id] = {"win": [], "loss": []}
        templates[sub_dim_id][result].append(tpl_text)

    wb.close()
    return templates


def _load_rank_average(xlsx_path: str) -> Dict[str, Dict[int, Dict]]:
    """
    从 rank_average.xlsx 的「段位平均水平」sheet 加载段位平均数据。

    Excel 列: 段位 | sub_dim_id | 指标名称 | 单位 | 平均表现
    返回: { "段位名": { sub_dim_id: { "unit": str, "avg": float } } }
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["段位平均水平"]

    rank_table: Dict[str, Dict[int, Dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rank_name = str(row[0]).strip()
        sub_dim_id = int(row[1])
        unit = str(row[3] or "")
        avg_val = row[4]
        if avg_val is None:
            continue
        avg_val = float(avg_val)

        if rank_name not in rank_table:
            rank_table[rank_name] = {}
        rank_table[rank_name][sub_dim_id] = {"unit": unit, "avg": avg_val}

    wb.close()
    return rank_table


# ================================================================
#  模块初始化：加载 Excel 数据（只在 import 时执行一次）
# ================================================================

SUB_DIM_DISPLAY_NAMES, SUB_DIM_UNITS, TIME_BASED_SUB_DIMS = _load_indicator_info(_TEMPLATES_XLSX)
HIGHLIGHT_TEMPLATES = _load_highlight_templates(_TEMPLATES_XLSX)
RANK_AVERAGE_TABLE = _load_rank_average(_RANK_AVG_XLSX)


# ================================================================
#  工具函数
# ================================================================

def format_raw_value(sub_dim_id: int, raw_value: int) -> str:
    """
    将子指标原始值格式化为展示用字符串。
    时间类指标从 ms 转为秒（保留1位小数），其余直接显示整数。
    """
    if sub_dim_id in TIME_BASED_SUB_DIMS:
        return f"{raw_value / 1000:.1f}"
    return str(raw_value)


# ================================================================
#  数据结构
# ================================================================

@dataclass
class HighlightReport:
    """高光点报告（三段式输出）"""
    player_gid: int
    is_win: bool
    player_rank: str
    match_id: str

    # 选中的最优指标信息
    best_dim: Optional[DimensionDelta] = None         # 优势维度 (max Δ)
    best_sub_dim_id: int = 0                           # 最优子指标 ID
    best_sub_dim_name: str = ""                        # 最优子指标展示名
    player_raw_value: int = 0                          # 玩家原始值
    opponent_raw_value: int = 0                        # 对手原始值
    player_formatted_value: str = ""                   # 玩家格式化值
    opponent_formatted_value: str = ""                 # 对手格式化值

    # 三段输出文本
    part1_highlight: str = ""                          # Part 1: 数据亮点
    part2_rank_average: str = ""                       # Part 2: 段位平均水平
    part3_detail_link: str = ""                        # Part 3: 详细数据链接
    detail_link_match_id: str = ""                     # 前端跳转用 match_id


# ================================================================
#  核心逻辑
# ================================================================

def _find_best_sub_indicator(
    advantage_dim: DimensionDelta,
    player_data: RadarChartData,
    opponent_data: RadarChartData,
) -> Optional[Tuple[SubDimensionData, Optional[SubDimensionData]]]:
    """
    在优势维度下，选取 max(玩家加权值 − 对手加权值) 的子指标。

    选取标准：玩家与对手在该子指标上的加权值差距最大的那个，
    即玩家相对优势最明显的子指标。

    Args:
        advantage_dim: 优势维度信息
        player_data:   玩家雷达图数据
        opponent_data: 对手雷达图数据

    Returns:
        (player_sub, opponent_sub) 或 None
    """
    player_dims = {d.dim_id: d for d in player_data.dimensions}
    opponent_dims = {d.dim_id: d for d in opponent_data.dimensions}

    player_dim = player_dims.get(advantage_dim.dim_id)
    if not player_dim or not player_dim.sub_dimensions:
        return None

    # 构建对手子指标映射
    opponent_dim = opponent_dims.get(advantage_dim.dim_id)
    opp_sub_map: Dict[int, SubDimensionData] = {}
    if opponent_dim:
        opp_sub_map = {s.sub_dim_id: s for s in opponent_dim.sub_dimensions}

    # 选取 max(玩家加权值 − 对手加权值) 的子指标
    best_player_sub: Optional[SubDimensionData] = None
    best_opponent_sub: Optional[SubDimensionData] = None
    max_delta = float('-inf')

    for p_sub in player_dim.sub_dimensions:
        o_sub = opp_sub_map.get(p_sub.sub_dim_id)
        o_weighted = o_sub.weighted_value if o_sub else 0
        delta = p_sub.weighted_value - o_weighted
        if delta > max_delta:
            max_delta = delta
            best_player_sub = p_sub
            best_opponent_sub = o_sub

    if best_player_sub is None:
        return None

    return (best_player_sub, best_opponent_sub)


def _render_part1_highlight(
    sub_dim_id: int,
    player_val_str: str,
    opponent_val_str: str,
    is_win: bool,
) -> str:
    """
    Part 1: 从模版库选取并渲染数据亮点文本。
    """
    templates = HIGHLIGHT_TEMPLATES.get(sub_dim_id)
    if not templates:
        display_name = SUB_DIM_DISPLAY_NAMES.get(sub_dim_id, f"指标{sub_dim_id}")
        return f"本场对战中您的{display_name}为{player_val_str}，做的很棒！"

    key = "win" if is_win else "loss"
    candidates = templates.get(key, [])
    if not candidates:
        candidates = templates.get("win", []) or templates.get("loss", [])

    template = random.choice(candidates)

    return template.format(
        player_val=player_val_str,
        opponent_val=opponent_val_str,
    )


def _render_part2_rank_average(
    sub_dim_id: int,
    player_rank: str,
) -> str:
    """
    Part 2: 根据用户段位与最佳子指标，生成段位平均水平文本。
    """
    rank_data = RANK_AVERAGE_TABLE.get(player_rank, {})
    indicator_avg = rank_data.get(sub_dim_id)

    display_name = SUB_DIM_DISPLAY_NAMES.get(sub_dim_id, f"指标{sub_dim_id}")

    if not indicator_avg:
        return f"全服玩家在段位【{player_rank}】的{display_name}平均数据暂无记录。"

    avg_val = indicator_avg["avg"]
    unit = indicator_avg["unit"]

    if isinstance(avg_val, float) and avg_val == int(avg_val) and avg_val >= 100:
        avg_display = str(int(avg_val))
    elif isinstance(avg_val, float):
        avg_display = f"{avg_val:.1f}"
    else:
        avg_display = str(avg_val)

    return f"全服玩家在段位【{player_rank}】，场均{avg_display}{unit}{display_name}"


def _render_part3_detail_link(match_id: str) -> str:
    """
    Part 3: 生成详细数据链接文本。
    """
    return f"查看本场战斗｜数据详情 ▶  [match_id: {match_id}]"


# ================================================================
#  报告生成与格式化
# ================================================================

def generate_highlight_report(
    player_gid: int,
    radar_data: RadarChartData,
    opponent_radar: RadarChartData,
    is_win: bool,
    player_rank: str = "万渊三",
    match_id: str = "MATCH_20260223_10001_20002",
) -> HighlightReport:
    """
    生成高光点报告（三段式输出）。

    核心流程:
      1. 计算维度分差 → 优势维度 = max(Δ_dimension)
      2. 在优势维度下 → 最优子指标 = max(玩家加权值)
      3. 根据子指标 + 胜负 → 选取模版生成三段输出

    Args:
        player_gid:     玩家 GID
        radar_data:     玩家雷达图数据
        opponent_radar: 对手雷达图数据
        is_win:         整局胜负（True=胜利, False=失败）
        player_rank:    玩家段位名称（用于段位平均查询）
        match_id:       对局 ID（前端跳转用）

    Returns:
        HighlightReport: 三段式高光点报告
    """
    report = HighlightReport(
        player_gid=player_gid,
        is_win=is_win,
        player_rank=player_rank,
        match_id=match_id,
        detail_link_match_id=match_id,
    )

    # Step 1: 统一计算维度分差
    delta_result = compute_dimension_deltas(radar_data, opponent_radar)
    adv = delta_result.advantage_dim
    report.best_dim = adv

    if not adv:
        report.part1_highlight = "本场对局暂无突出的数据亮点。"
        report.part3_detail_link = _render_part3_detail_link(match_id)
        return report

    # Step 2: 在优势维度下选取最优子指标
    result = _find_best_sub_indicator(adv, radar_data, opponent_radar)
    if not result:
        report.part1_highlight = "本场对局暂无突出的数据亮点。"
        report.part3_detail_link = _render_part3_detail_link(match_id)
        return report

    player_sub, opponent_sub = result
    sub_dim_id = player_sub.sub_dim_id

    # 记录选中指标信息
    report.best_sub_dim_id = sub_dim_id
    report.best_sub_dim_name = SUB_DIM_DISPLAY_NAMES.get(sub_dim_id, player_sub.name)
    report.player_raw_value = player_sub.raw_value
    report.opponent_raw_value = opponent_sub.raw_value if opponent_sub else 0

    # 格式化原始值
    p_val = format_raw_value(sub_dim_id, player_sub.raw_value)
    o_val = format_raw_value(sub_dim_id, opponent_sub.raw_value) if opponent_sub else "0"
    report.player_formatted_value = p_val
    report.opponent_formatted_value = o_val

    # Step 3: 三段式输出
    report.part1_highlight = _render_part1_highlight(sub_dim_id, p_val, o_val, is_win)
    report.part2_rank_average = _render_part2_rank_average(sub_dim_id, player_rank)
    report.part3_detail_link = _render_part3_detail_link(match_id)

    return report


def format_highlight_report(report: HighlightReport) -> str:
    """
    格式化高光点报告为可读文本（三段式）。

    设计原则：根据加权值选定指标，但不展示加权值/维度计算分，
    面向用户只展示真实指标值。
    """
    win_label = "胜利" if report.is_win else "失败"
    lines = [f"===== 玩家 {report.player_gid} 高光点 ({win_label}) =====\n"]

    # Part 1: 数据亮点（模版文案，不额外展示指标名+数值的标题行）
    lines.append(f"⭐ {report.part1_highlight}")
    lines.append("")

    # Part 2: 段位平均水平
    if report.part2_rank_average:
        lines.append(f"📊 {report.part2_rank_average}")
        lines.append("")

    # Part 3: 详细数据链接
    lines.append(f"🔗 {report.part3_detail_link}")

    return "\n".join(lines)
