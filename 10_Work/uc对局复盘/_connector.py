# -*- coding: utf-8 -*-
"""为文案模版+建议添加承接词"""
import openpyxl
import re

wb = openpyxl.load_workbook('improvement_templates.xlsx')
ws = wb['文案x建议']

def add_connector(combined: str) -> str:
    """在文案模版和建议之间添加承接词"""
    parts = combined.split('\n', 1)
    if len(parts) < 2:
        return combined
    template = parts[0].strip()
    suggestion = parts[1].strip()
    if not template or not suggestion:
        return combined

    # 去掉模版末尾句号，统一处理
    tpl = template.rstrip('。')

    # 分析建议开头，选择承接方式
    # 类型1: "值偏低/偏高提醒..." -> 去掉前缀，转为建议句
    m = re.match(r'^值偏[低高]提醒(.+)', suggestion)
    if m:
        return f'{tpl}，{m.group(1)}'

    # 类型2: "与XX做比..." / "综合...分析" / "结合...分析" -> 分析说明型，用句号隔开
    if re.match(r'^(与.+做比|综合.+分析|结合.+分析|过高代表)', suggestion):
        return f'{tpl}。{suggestion}'

    # 类型3: 直接建议/动作型 -> 加"建议"承接
    # 但如果建议本身以"建议"开头则不重复
    if suggestion.startswith('建议'):
        return f'{tpl}，{suggestion}'

    return f'{tpl}，建议{suggestion}'


changed = 0
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=3).value
    if val:
        new_val = add_connector(val)
        if new_val != val:
            ws.cell(row=r, column=3, value=new_val)
            changed += 1

wb.save('improvement_templates.xlsx')
print(f'Done! Updated {changed} rows.')
