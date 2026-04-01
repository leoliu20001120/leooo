# -*- coding: utf-8 -*-
"""
Excel知识库生成器
使用openpyxl创建多Sheet的格式化Excel文件
"""
import logging
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config

logger = logging.getLogger("ExcelGenerator")


class ExcelGenerator:
    """Excel知识库生成器"""

    # 样式定义
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DATA_FONT = Font(name="微软雅黑", size=10)
    DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 等级颜色
    RARITY_FILLS = {
        "白银": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
        "黄金": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "棱彩": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }

    # 推荐等级颜色
    GRADE_FILLS = {
        "S": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "A": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
        "B": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "C": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        "D": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    }

    GRADE_FONTS = {
        "S": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
        "A": Font(name="微软雅黑", size=11, bold=True, color="000000"),
        "B": Font(name="微软雅黑", size=11, bold=True, color="000000"),
        "C": Font(name="微软雅黑", size=10, color="000000"),
        "D": Font(name="微软雅黑", size=10, color="666666"),
    }

    def __init__(self):
        self.wb = Workbook()

    def generate(self, merged_data, output_path=None):
        """生成完整的Excel知识库"""
        output_path = output_path or config.FINAL_EXCEL_PATH
        logger.info(f"开始生成Excel知识库: {output_path}")

        augments = merged_data.get("augments", [])
        combos = merged_data.get("combos", [])
        sets = merged_data.get("sets", [])
        champion_combos = merged_data.get("champion_combos", [])
        arammayhem_combos = merged_data.get("arammayhem_combos", [])

        # Sheet1: 符文基础信息表（核心表）
        self._create_augment_sheet(augments)

        # Sheet2: 符文UGC评论表
        self._create_ugc_sheet(augments)

        # Sheet3: 英雄推荐海克斯组合（从hextech.dtodo.cn爬取）
        self._create_champion_combo_sheet(champion_combos)

        # Sheet4: 英雄符文搭配（从arammayhem.com爬取）
        self._create_arammayhem_sheet(arammayhem_combos)

        # Sheet5: 符文推荐理由表
        self._create_recommend_sheet(augments)

        # Sheet6: 符文套装系统表
        self._create_sets_sheet(sets)

        # Sheet6: 符文冷知识表
        self._create_funfact_sheet(augments)

        # Sheet7: 符文适配英雄表
        self._create_champion_sheet(augments)

        # 删除默认sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # 保存
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.wb.save(output_path)
        logger.info(f"Excel知识库生成完成: {output_path}")
        return output_path

    def _apply_header_style(self, ws, row, col_count):
        """应用表头样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER

    def _apply_data_style(self, ws, row, col_count):
        """应用数据行样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.DATA_FONT
            cell.alignment = self.DATA_ALIGN
            cell.border = self.THIN_BORDER

    def _auto_width(self, ws, col_widths):
        """设置列宽"""
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # ===== Sheet1: 符文基础信息表 =====
    def _create_augment_sheet(self, augments):
        ws = self.wb.create_sheet("符文基础信息", 0)

        # 表头
        headers = [
            "序号", "符文名称", "等级", "Tier分级", "推荐指数", "推荐icon",
            "胜率(%)", "选取率(%)", "官方描述（掌盟）", "第三方描述（hextech）",
            "玩家补充描述", "人话描述",
            "UGC评分", "评分样本数", "评论数", "所属套装",
            "是否新符文", "icon_URL", "审核状态"
        ]
        col_widths = [6, 18, 8, 8, 8, 10, 10, 10, 50, 50, 50, 50, 10, 10, 10, 20, 10, 40, 10]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 数据
        for idx, aug in enumerate(augments, 1):
            row = idx + 1
            grade = aug.get("recommendation_grade", "C")
            related_sets = aug.get("related_sets", [])

            # 评分直接来自掌盟 go/vote/get_rate 接口，是真实的10分制平均分
            ugc_score = aug.get("ugc_score", 0)
            score_count = aug.get("ugc_score_count", 0)

            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=aug.get("name", ""))
            ws.cell(row=row, column=3, value=aug.get("rarity", ""))
            ws.cell(row=row, column=4, value=aug.get("tier", ""))
            ws.cell(row=row, column=5, value=grade)
            ws.cell(row=row, column=6, value=aug.get("recommendation_icon", ""))
            ws.cell(row=row, column=7, value=aug.get("win_rate", 0))
            ws.cell(row=row, column=8, value=aug.get("pick_rate", 0))
            ws.cell(row=row, column=9, value=aug.get("zhangmeng_desc", ""))
            ws.cell(row=row, column=10, value=aug.get("hextech_desc", ""))
            ws.cell(row=row, column=11, value=aug.get("tooltip_desc", ""))
            ws.cell(row=row, column=12, value=aug.get("plain_desc", ""))
            ws.cell(row=row, column=13, value=ugc_score)
            ws.cell(row=row, column=14, value=score_count)
            ws.cell(row=row, column=15, value=aug.get("ugc_total_comments", 0))
            ws.cell(row=row, column=16, value="、".join(related_sets) if related_sets else "")
            ws.cell(row=row, column=17, value="是" if aug.get("is_new") else "否")
            ws.cell(row=row, column=18, value=aug.get("icon_url", ""))
            ws.cell(row=row, column=19, value="待审核")

            self._apply_data_style(ws, row, len(headers))

            # 等级颜色
            rarity = aug.get("rarity", "")
            if rarity in self.RARITY_FILLS:
                ws.cell(row=row, column=3).fill = self.RARITY_FILLS[rarity]

            # 推荐指数颜色
            if grade in self.GRADE_FILLS:
                ws.cell(row=row, column=5).fill = self.GRADE_FILLS[grade]
                ws.cell(row=row, column=5).font = self.GRADE_FONTS.get(grade, self.DATA_FONT)
                ws.cell(row=row, column=5).alignment = self.CENTER_ALIGN

            # 居中列
            for c in [1, 3, 4, 5, 6, 7, 8, 13, 14, 15, 17, 19]:
                ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN

        # 自动筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(augments) + 1}"

    # ===== Sheet2: 符文UGC评论表 =====
    def _create_ugc_sheet(self, augments):
        ws = self.wb.create_sheet("符文UGC评论")

        headers = ["符文名称", "评论内容", "来源地区", "评分", "点赞数", "来源"]
        col_widths = [18, 60, 15, 8, 10, 10]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        row = 2
        for aug in augments:
            name = aug.get("name", "")
            comments = aug.get("ugc_comments", [])
            for comment in comments:
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=comment.get("content", ""))
                # 来源地区（from_addr，如"来自四川"）
                ws.cell(row=row, column=3, value=comment.get("from_addr", ""))
                # 该评论者的评分（1-5分，0=未评分）
                rate = comment.get("rate", 0)
                ws.cell(row=row, column=4, value=rate if rate > 0 else "")
                ws.cell(row=row, column=5, value=comment.get("likes", 0))
                ws.cell(row=row, column=6, value="掌盟")
                self._apply_data_style(ws, row, len(headers))
                # 居中对齐评分和点赞列
                for c in [3, 4, 5, 6]:
                    ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN
                row += 1

        if row == 2:
            ws.cell(row=2, column=1, value="暂无评论数据（需要通过掌盟APP抓包获取）")

    # ===== Sheet3: 英雄推荐海克斯组合表（从hextech.dtodo.cn爬取） =====
    def _create_champion_combo_sheet(self, champion_combos):
        ws = self.wb.create_sheet("英雄推荐海克斯组合")

        headers = [
            "英雄名称", "英雄称号", "英雄层级", "英雄胜率", "英雄选取率",
            "组合排名", "组合符文", "组合层级"
        ]
        col_widths = [14, 14, 8, 10, 10, 8, 50, 8]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        # 层级颜色
        TIER_FILLS = {
            "T1": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "T2": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
            "T3": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "T4": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            "T5": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
        }
        TIER_FONTS = {
            "T1": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
            "T2": Font(name="微软雅黑", size=10, bold=True, color="000000"),
            "T3": Font(name="微软雅黑", size=10, bold=True, color="000000"),
            "T4": Font(name="微软雅黑", size=10, color="000000"),
            "T5": Font(name="微软雅黑", size=10, color="666666"),
        }

        row = 2
        champion_count = 0
        combo_count = 0

        # 按英雄层级和胜率排序（T1最前，同层级按胜率降序）
        tier_order = {"T1": 1, "T2": 2, "T3": 3, "T4": 4, "T5": 5, "": 9}
        sorted_champions = sorted(
            champion_combos,
            key=lambda x: (
                tier_order.get(x.get("tier", ""), 9),
                -float(x.get("win_rate", "0%").replace("%", "") or 0)
            )
        )

        for champ in sorted_champions:
            combos = champ.get("combos", [])
            if not combos:
                continue  # 跳过没有组合数据的英雄

            champion_count += 1
            champion_name = champ.get("champion_name", "")
            champion_title = champ.get("champion_title", "")
            champion_tier = champ.get("tier", "")
            win_rate = champ.get("win_rate", "")
            pick_rate = champ.get("pick_rate", "")

            for combo in combos:
                combo_rank = combo.get("rank", 0)
                augments_list = combo.get("augments", [])
                combo_tier = combo.get("tier", "")
                augments_str = " + ".join(augments_list)

                ws.cell(row=row, column=1, value=champion_name)
                ws.cell(row=row, column=2, value=champion_title)
                ws.cell(row=row, column=3, value=champion_tier)
                ws.cell(row=row, column=4, value=win_rate)
                ws.cell(row=row, column=5, value=pick_rate)
                ws.cell(row=row, column=6, value=combo_rank)
                ws.cell(row=row, column=7, value=augments_str)
                ws.cell(row=row, column=8, value=combo_tier)

                self._apply_data_style(ws, row, len(headers))

                # 居中列
                for c in [1, 2, 3, 4, 5, 6, 8]:
                    ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN

                # 英雄层级颜色
                if champion_tier in TIER_FILLS:
                    ws.cell(row=row, column=3).fill = TIER_FILLS[champion_tier]
                    ws.cell(row=row, column=3).font = TIER_FONTS.get(champion_tier, self.DATA_FONT)

                # 组合层级颜色
                if combo_tier in TIER_FILLS:
                    ws.cell(row=row, column=8).fill = TIER_FILLS[combo_tier]
                    ws.cell(row=row, column=8).font = TIER_FONTS.get(combo_tier, self.DATA_FONT)

                row += 1
                combo_count += 1

        # 自动筛选
        if row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

        logger.info(f"英雄推荐海克斯组合: {champion_count}个英雄, {combo_count}条组合数据")

    # ===== Sheet4: 英雄符文搭配（从arammayhem.com爬取） =====
    # 手动补充的黑科技玩法数据（拆解自社区攻略）
    MANUAL_SUPPLEMENT_COMBOS = [
        # 1. 体型流
        {"champion_name": "（坦克通用）", "augment_name": "坦克引擎", "tier": "S", "tags": ["黑科技"], "description": "【体型流】击杀叠层增体型和生命值，选择坦克英雄优先堆叠心之钢层数，配合歌利亚巨人和任务：钢化你心，变得又肉又有输出，成为全场最大的存在", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（坦克通用）", "augment_name": "歌利亚巨人", "tier": "S", "tags": ["黑科技"], "description": "【体型流】直接增体型、生命值和适应之力，配合坦克引擎和任务：钢化你心堆叠心之钢层数，让你在战场上极具威慑力", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（坦克通用）", "augment_name": "任务：钢化你心", "tier": "S", "tags": ["黑科技"], "description": "【体型流】持有心之钢并达到一定层数后将层数翻倍，配合坦克引擎和歌利亚巨人，体型和生命值无限膨胀", "upvotes": 0, "downvotes": 0, "score": 0},
        # 2. 踢踏舞流
        {"champion_name": "（高攻速通用）", "augment_name": "踢踏舞", "tier": "A", "tags": ["黑科技"], "description": "【踢踏舞流】攻击叠加可无限叠加的移动速度和攻击速度，适合高攻速英雄，通过持续攻击保持高速移动，闪电般穿梭战场，体验飞一般的感觉", "upvotes": 0, "downvotes": 0, "score": 0},
        # 3. 冰寒火男
        {"champion_name": "布兰德", "augment_name": "炼狱导管", "tier": "S", "tags": ["黑科技"], "description": "【冰寒火男】技能施加无限叠加的灼烧并减CD，布兰德技能全是AOE，配合冰寒减速和魔法飞弹弹射，出冰杖等减速装备持续减速并灼烧敌人，打出成吨伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "布兰德", "augment_name": "魔法飞弹", "tier": "A", "tags": ["黑科技"], "description": "【冰寒火男】技能伤害触发真实伤害飞弹，配合布兰德AOE技能和炼狱导管灼烧+冰寒减速，冰火两重天全屏控制", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "布兰德", "augment_name": "冰寒", "tier": "A", "tags": ["黑科技"], "description": "【冰寒火男】强化减速效果，配合布兰德AOE技能+炼狱导管灼烧+魔法飞弹，让敌人跑不掉持续吃持续伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        # 4. 一板一眼杰斯
        {"champion_name": "杰斯", "augment_name": "一板一眼", "tier": "S", "tags": ["黑科技"], "description": "【一板一眼杰斯】攻击速度固定，额外攻速转化为攻击力。杰斯炮形态下开启W技能获得大量额外攻速，通过符文转化为高额攻击力，QE二连伤害极高，瞬间融化对手", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "盖伦", "augment_name": "一板一眼", "tier": "A", "tags": ["黑科技"], "description": "【一板一眼】攻速固定转化为攻击力，盖伦E旋转连续命中后接Q单击天然交替攻击触发一板一眼，也非常适配此符文", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "卡特琳娜", "augment_name": "一板一眼", "tier": "A", "tags": ["黑科技"], "description": "【一板一眼】攻速固定转化为攻击力，卡特琳娜等依赖攻速的英雄也可使用此符文获得极致爆发伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        # 5. 电风扇人马
        {"champion_name": "赫卡里姆", "augment_name": "虚幻武器", "tier": "S", "tags": ["黑科技"], "description": "【电风扇人马】技能可施加攻击特效（每个目标有1秒冷却），人马利用Q技能持续伤害触发虚幻武器特效，配合秘术冲拳频繁刷新技能+W技能恢复，在人群中持续作战", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "赫卡里姆", "augment_name": "秘术冲拳", "tier": "S", "tags": ["黑科技"], "description": "【电风扇人马】技能命中减基础技能CD，人马配合虚幻武器让Q技能持续触发攻击特效，频繁刷新技能实现高频率输出和高额恢复", "upvotes": 0, "downvotes": 0, "score": 0},
        # 6. 瑞兹无限循环流
        {"champion_name": "瑞兹", "augment_name": "由心即物", "tier": "S", "tags": ["黑科技"], "description": "【瑞兹无限循环流】蓝量转化为生命值，配合物理转魔法（攻击力转法强）+霸王血铠（生命值转攻击力）+魔切（提升蓝量），形成蓝量→生命值→攻击力→法强→蓝量的无限循环，后期一个Q即可造成巨额伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "瑞兹", "augment_name": "物理转魔法", "tier": "S", "tags": ["黑科技"], "description": "【瑞兹无限循环流】攻击力转化为法强并百分比提升，配合由心即物和霸王血铠+魔切形成属性闭环，实现法强和生命值的无限增长", "upvotes": 0, "downvotes": 0, "score": 0},
        # 7. 剑圣无限Q
        {"champion_name": "易", "augment_name": "秘术冲拳", "tier": "S", "tags": ["黑科技"], "description": "【剑圣无限Q】技能命中减基础技能CD，搭配循环往复（+60技能急速）或面包黄油（+100技能急速），并出吸蓝刀、青龙刀等装备使Q技能急速达到190以上，即可实现无限Q的效果", "upvotes": 0, "downvotes": 0, "score": 0},
        # 8. 虚幻武器特效流
        {"champion_name": "布兰德", "augment_name": "虚幻武器", "tier": "A", "tags": ["黑科技"], "description": "【虚幻武器特效流】技能可施加攻击特效（每个目标有1秒冷却），布兰德有持续伤害技能，可选择特效流装备（如破败王者之刃）打伤害，或用心之钢快速叠加生命值", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "卡西奥佩娅", "augment_name": "虚幻武器", "tier": "A", "tags": ["黑科技"], "description": "【虚幻武器特效流】技能可施加攻击特效，蛇女E技能高频命中可持续触发特效，配合特效流装备或心之钢快速叠加", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "阿木木", "augment_name": "虚幻武器", "tier": "A", "tags": ["黑科技"], "description": "【虚幻武器特效流】技能可施加攻击特效，阿木木W持续伤害技能可持续触发特效，配合心之钢快速叠加生命值", "upvotes": 0, "downvotes": 0, "score": 0},
        # 9. AP暴击流
        {"champion_name": "布兰德", "augment_name": "珠光护手", "tier": "A", "tags": ["黑科技"], "description": "【AP暴击流】技能可暴击，配合易损（装备和持续伤害可暴击）和关键暴击（提供高额暴击率），出无尽之刃+灭世者的死亡之帽，技能暴击伤害极高", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "瑞兹", "augment_name": "珠光护手", "tier": "A", "tags": ["黑科技"], "description": "【AP暴击流】技能可暴击，瑞兹高频Q技能配合易损+关键暴击，出暴击装+法强装，每个Q都有暴击可能，伤害爆炸", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（法师通用）", "augment_name": "易损", "tier": "A", "tags": ["黑科技"], "description": "【AP暴击流】装备和持续伤害可暴击，配合珠光护手（技能可暴击）和关键暴击（高额暴击率），让法师的技能也能造成暴击", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（法师通用）", "augment_name": "关键暴击", "tier": "A", "tags": ["黑科技"], "description": "【AP暴击流】提供高额暴击率，配合珠光护手+易损让法师技能暴击，出无尽之刃瞬间打出高额伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        # 10. 其他趣味玩法
        {"champion_name": "普朗克", "augment_name": "回归基本功", "tier": "A", "tags": ["黑科技"], "description": "【回归基本功船长】封印大招但大幅增强技能伤害和技能急速，船长的每个桶都算技能，可频繁释放造成爆炸伤害", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "菲兹", "augment_name": "全凭身法", "tier": "A", "tags": ["黑科技"], "description": "【全凭身法小鱼人】冲刺、跳跃类技能获得技能急速，让小鱼人的QE技能冷却极短，灵活穿梭战场", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（通用）", "augment_name": "闪现向前", "tier": "B", "tags": ["黑科技"], "description": "【无限闪现】闪现向前和闪闪现现组合，获得多个闪现，创造无限可能，适合需要灵活位移的英雄", "upvotes": 0, "downvotes": 0, "score": 0},
        {"champion_name": "（通用）", "augment_name": "闪闪现现", "tier": "B", "tags": ["黑科技"], "description": "【无限闪现】配合闪现向前获得多个闪现，创造无限位移可能，适合需要灵活走位的各类英雄", "upvotes": 0, "downvotes": 0, "score": 0},
    ]

    def _create_arammayhem_sheet(self, arammayhem_combos):
        ws = self.wb.create_sheet("英雄符文搭配")

        headers = [
            "英雄名称", "符文名称", "评级", "标签", "描述",
            "点赞", "点踩", "分数"
        ]
        col_widths = [14, 18, 6, 14, 70, 8, 8, 8]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        # 评级颜色
        TIER_FILLS = {
            "S": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "A": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
            "B": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "C": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        }
        TIER_FONTS = {
            "S": Font(name="微软雅黑", size=10, bold=True, color="FFFFFF"),
            "A": Font(name="微软雅黑", size=10, bold=True, color="000000"),
            "B": Font(name="微软雅黑", size=10, bold=True, color="000000"),
            "C": Font(name="微软雅黑", size=10, color="000000"),
        }

        # 标签颜色
        TAG_FILLS = {
            "神级": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "强力": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "娱乐": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            "黑科技": PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid"),
            "陷阱": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        }
        TAG_FONTS = {
            "神级": Font(name="微软雅黑", size=10, bold=True, color="996600"),
            "强力": Font(name="微软雅黑", size=10, bold=True, color="006600"),
            "娱乐": Font(name="微软雅黑", size=10, color="003366"),
            "黑科技": Font(name="微软雅黑", size=10, bold=True, color="7B2D8E"),
            "陷阱": Font(name="微软雅黑", size=10, color="CC0000"),
        }

        # 手动补充行的底色（浅黄绿色，用于区分）
        SUPPLEMENT_FILL = PatternFill(
            start_color="FFFFED", end_color="FFFFED", fill_type="solid"
        )

        if not arammayhem_combos:
            arammayhem_combos = []

        # 按英雄分组，英雄内按分数降序、评级排序
        tier_order = {"S": 1, "A": 2, "B": 3, "C": 4, "": 9}
        sorted_combos = sorted(
            arammayhem_combos,
            key=lambda x: (
                x.get("champion_name", ""),
                tier_order.get(x.get("tier", ""), 9),
                -x.get("score", 0),
            )
        )

        row = 2

        def _write_combo_row(ws, row, combo, headers, is_supplement=False):
            """写入一行combo数据"""
            champion_name = combo.get("champion_name", "")
            augment_name = combo.get("augment_name", "")
            tier = combo.get("tier", "")
            tags = combo.get("tags", [])
            description = combo.get("description", "")
            upvotes = combo.get("upvotes", 0)
            downvotes = combo.get("downvotes", 0)
            score = combo.get("score", 0)

            tags_str = "、".join(tags) if isinstance(tags, list) else str(tags)

            ws.cell(row=row, column=1, value=champion_name)
            ws.cell(row=row, column=2, value=augment_name)
            ws.cell(row=row, column=3, value=tier)
            ws.cell(row=row, column=4, value=tags_str)
            ws.cell(row=row, column=5, value=description)
            ws.cell(row=row, column=6, value=upvotes)
            ws.cell(row=row, column=7, value=downvotes)
            ws.cell(row=row, column=8, value=score)

            self._apply_data_style(ws, row, len(headers))

            # 手动补充行加底色
            if is_supplement:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = SUPPLEMENT_FILL

            # 居中列
            for c in [1, 2, 3, 4, 6, 7, 8]:
                ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN

            # 描述列自动换行
            ws.cell(row=row, column=5).alignment = Alignment(
                vertical="center", wrap_text=True
            )

            # 评级颜色（补充行底色不覆盖评级色）
            if tier in TIER_FILLS:
                ws.cell(row=row, column=3).fill = TIER_FILLS[tier]
                ws.cell(row=row, column=3).font = TIER_FONTS.get(tier, self.DATA_FONT)

            # 标签颜色
            if tags:
                first_tag = tags[0] if isinstance(tags, list) else str(tags)
                if first_tag in TAG_FILLS:
                    ws.cell(row=row, column=4).fill = TAG_FILLS[first_tag]
                    ws.cell(row=row, column=4).font = TAG_FONTS.get(first_tag, self.DATA_FONT)

            # 分数正值绿色、负值红色
            if score > 0:
                ws.cell(row=row, column=8).font = Font(
                    name="微软雅黑", size=10, bold=True, color="006600"
                )
            elif score < 0:
                ws.cell(row=row, column=8).font = Font(
                    name="微软雅黑", size=10, bold=True, color="CC0000"
                )

        # 写入爬取的数据
        for combo in sorted_combos:
            _write_combo_row(ws, row, combo, headers, is_supplement=False)
            row += 1

        # 在末尾追加手动补充的黑科技玩法数据（带底色标识）
        supplement_start_row = row
        for combo in self.MANUAL_SUPPLEMENT_COMBOS:
            _write_combo_row(ws, row, combo, headers, is_supplement=True)
            row += 1

        supplement_count = len(self.MANUAL_SUPPLEMENT_COMBOS)

        # 自动筛选
        if row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

        # 统计
        champion_count = len(set(c.get("champion_name", "") for c in arammayhem_combos))
        logger.info(
            f"英雄符文搭配: {champion_count}个英雄, {len(arammayhem_combos)}条爬取数据 "
            f"+ {supplement_count}条手动补充黑科技玩法（第{supplement_start_row}-{row-1}行，浅黄底色）"
        )

    # ===== Sheet4: 符文推荐理由表 =====
    def _create_recommend_sheet(self, augments):
        ws = self.wb.create_sheet("符文推荐理由")

        headers = ["符文名称", "推荐tag", "短评（5-10字）", "适配英雄类型", "推荐指数", "审核状态"]
        col_widths = [18, 20, 30, 25, 10, 10]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        for idx, aug in enumerate(augments, 1):
            row = idx + 1
            types = aug.get("champion_types", [])
            grade = aug.get("recommendation_grade", "")

            ws.cell(row=row, column=1, value=aug.get("name", ""))
            ws.cell(row=row, column=2, value=aug.get("recommend_tag", ""))
            ws.cell(row=row, column=3, value=aug.get("short_review", ""))
            ws.cell(row=row, column=4, value="、".join(types) if types else "")
            ws.cell(row=row, column=5, value=grade)
            ws.cell(row=row, column=6, value="待审核")
            self._apply_data_style(ws, row, len(headers))

            if grade in self.GRADE_FILLS:
                ws.cell(row=row, column=5).fill = self.GRADE_FILLS[grade]
                ws.cell(row=row, column=5).alignment = self.CENTER_ALIGN

    # ===== Sheet5: 符文套装/羁绊系统表（从hextech.dtodo.cn/synergy爬取） =====
    def _create_sets_sheet(self, sets):
        ws = self.wb.create_sheet("套装羁绊系统")

        headers = [
            "套装名称", "套装评级", "套装效果",
            "层级效果", "羁绊内海克斯", "海克斯数量",
            "策略建议", "组合技巧"
        ]
        col_widths = [16, 10, 50, 50, 50, 10, 60, 60]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        # 套装评级颜色
        SET_TIER_FILLS = {
            "S": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "A": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
            "B": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "C": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        }
        SET_TIER_FONTS = {
            "S": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
            "A": Font(name="微软雅黑", size=11, bold=True, color="000000"),
            "B": Font(name="微软雅黑", size=11, bold=True, color="000000"),
            "C": Font(name="微软雅黑", size=10, color="000000"),
        }

        # 按评级排序: S > A > B > C
        tier_order = {"S": 1, "A": 2, "B": 3, "C": 4, "D": 5, "": 9}
        sorted_sets = sorted(sets, key=lambda x: tier_order.get(x.get("tier", ""), 9))

        for idx, s in enumerate(sorted_sets, 1):
            row = idx + 1
            name = s.get("name", "")
            tier = s.get("tier", "")
            effect = s.get("effect", "")

            # 构建层级效果字符串
            tier_effects = s.get("tier_effects", {})
            if tier_effects and isinstance(tier_effects, dict):
                tier_effect_lines = []
                for k in sorted(tier_effects.keys(), key=lambda x: int(x.replace("件", "")) if x.replace("件", "").isdigit() else 99):
                    tier_effect_lines.append(f"{k}: {tier_effects[k]}")
                tier_effect_str = "\n".join(tier_effect_lines)
            else:
                tier_effect_str = s.get("detailed_effect", "")

            # 构建羁绊内海克斯字符串
            augs = s.get("augments", [])
            augs_str = "、".join(augs) if isinstance(augs, list) else str(augs)

            # 海克斯数量
            aug_count = s.get("augment_count", len(augs) if isinstance(augs, list) else 0)

            # 策略建议和组合技巧
            strategy = s.get("strategy", "")
            combo_tips = s.get("combo_tips", "")

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=tier)
            ws.cell(row=row, column=3, value=effect)
            ws.cell(row=row, column=4, value=tier_effect_str)
            ws.cell(row=row, column=5, value=augs_str)
            ws.cell(row=row, column=6, value=aug_count)
            ws.cell(row=row, column=7, value=strategy)
            ws.cell(row=row, column=8, value=combo_tips)
            self._apply_data_style(ws, row, len(headers))

            # 套装评级颜色
            if tier in SET_TIER_FILLS:
                ws.cell(row=row, column=2).fill = SET_TIER_FILLS[tier]
                ws.cell(row=row, column=2).font = SET_TIER_FONTS.get(tier, self.DATA_FONT)

            # 居中列
            for c in [1, 2, 6]:
                ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN

            # 自动换行列
            for c in [3, 4, 5, 7, 8]:
                ws.cell(row=row, column=c).alignment = Alignment(vertical="center", wrap_text=True)

        logger.info(f"套装羁绊系统: {len(sorted_sets)}个套装（含策略建议+组合技巧，数据来源: hextech.dtodo.cn/synergy）")

    # ===== Sheet6: 符文冷知识表 =====
    def _create_funfact_sheet(self, augments):
        ws = self.wb.create_sheet("符文冷知识")

        headers = ["符文名称", "冷知识内容", "类型", "审核状态"]
        col_widths = [18, 80, 12, 10]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        row = 2
        for aug in augments:
            fun_fact = aug.get("fun_fact", "")
            if fun_fact:
                ws.cell(row=row, column=1, value=aug.get("name", ""))
                ws.cell(row=row, column=2, value=fun_fact)

                # 判断类型
                content = fun_fact.lower()
                if "套装" in content or "组合" in content or "搭配" in content:
                    fact_type = "组合类"
                elif "热评" in content or "评论" in content:
                    fact_type = "UGC类"
                elif "被动" in content or "效果" in content or "触发" in content:
                    fact_type = "机制类"
                else:
                    fact_type = "趣味类"

                ws.cell(row=row, column=3, value=fact_type)
                ws.cell(row=row, column=4, value="待审核")
                self._apply_data_style(ws, row, len(headers))
                row += 1

    # ===== Sheet7: 符文适配英雄表 =====
    def _create_champion_sheet(self, augments):
        ws = self.wb.create_sheet("符文适配英雄")

        headers = ["符文名称", "英雄名称", "排名"]
        col_widths = [18, 18, 8]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        self._auto_width(ws, col_widths)
        ws.freeze_panes = "A2"

        row = 2
        for aug in augments:
            name = aug.get("name", "")
            champions = aug.get("top_champions", [])
            for rank, champ in enumerate(champions, 1):
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=champ)
                ws.cell(row=row, column=3, value=rank)
                self._apply_data_style(ws, row, len(headers))
                for c in [1, 2, 3]:
                    ws.cell(row=row, column=c).alignment = self.CENTER_ALIGN
                row += 1
