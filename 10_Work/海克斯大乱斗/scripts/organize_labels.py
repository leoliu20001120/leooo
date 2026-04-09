"""
整理预选池符文标签：将最佳拍档/强力单卡/娱乐三个页签的标签统一合并到预选池
输出一个新的Excel文档，包含多个页签
"""
import pandas as pd
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 路径
base = '/Users/liusixing_tx/Documents/Obsidian Vault/10_Work/海克斯大乱斗'
filepath = f'{base}/测试.xlsx'
hero_map_path = f'{base}/data/英雄id定位表.xlsx'
output_path = f'{base}/预选池符文标签整理.xlsx'

# ============================================================
# 读取数据
# ============================================================
df_pool = pd.read_excel(filepath, sheet_name='完整预选池')
df_best = pd.read_excel(filepath, sheet_name='最佳拍档')
df_strong = pd.read_excel(filepath, sheet_name='强力单卡')
df_fun = pd.read_excel(filepath, sheet_name='娱乐')
df_hero_map = pd.read_excel(hero_map_path)

# 建立映射
title_to_name = dict(zip(df_hero_map['称号'], df_hero_map['中文名']))
name_to_title = dict(zip(df_hero_map['中文名'], df_hero_map['称号']))
title_to_role1 = dict(zip(df_hero_map['称号'], df_hero_map['定位1']))
title_to_role2 = dict(zip(df_hero_map['称号'], df_hero_map['定位2']))

# ============================================================
# 构建三个标签页签的 (称号, 符文名) 集合
# ============================================================
best_pairs = set(zip(df_best['英雄'], df_best['符文']))
strong_pairs = set(zip(df_strong['champion_name'], df_strong['augment_name']))

df_fun['称号'] = df_fun['英雄ID'].map(name_to_title)
fun_valid = df_fun.dropna(subset=['称号'])
fun_pairs = set(zip(fun_valid['称号'], fun_valid['符文ID']))

print(f'最佳拍档条目: {len(best_pairs)}')
print(f'强力单卡条目: {len(strong_pairs)}')
print(f'娱乐条目: {len(fun_pairs)}')

# ============================================================
# 为预选池每一行打标签
# ============================================================
df = df_pool.copy()

df['是否最佳拍档'] = df.apply(lambda r: (r['英雄名称'], r['符文名称']) in best_pairs, axis=1)
df['是否强力单卡'] = df.apply(lambda r: (r['英雄名称'], r['符文名称']) in strong_pairs, axis=1)
df['是否娱乐'] = df.apply(lambda r: (r['英雄名称'], r['符文名称']) in fun_pairs, axis=1)

def get_combined_label(row):
    labels = []
    if row['是否最佳拍档']: labels.append('最佳拍档')
    if row['是否强力单卡']: labels.append('强力单卡')
    if row['是否娱乐']: labels.append('娱乐')
    return '、'.join(labels) if labels else '无'

def get_priority_label(row):
    if row['是否最佳拍档']: return '最佳拍档'
    elif row['是否娱乐']: return '娱乐'
    elif row['是否强力单卡']: return '强力单卡'
    else: return '无'

df['符文标签'] = df.apply(get_combined_label, axis=1)
df['优先级标签'] = df.apply(get_priority_label, axis=1)
df['英雄中文名'] = df['英雄名称'].map(title_to_name)
df['定位1'] = df['英雄名称'].map(title_to_role1)
df['定位2'] = df['英雄名称'].map(title_to_role2).fillna('')

# ============================================================
# 整理各表
# ============================================================

# 完整表
df_full = df[[
    '英雄名称', '英雄中文名', '定位1', '定位2',
    '符文名称', '符文ID', '稀有度', '评级', '表现分', '热度',
    '标签', '符文标签', '优先级标签',
    '是否最佳拍档', '是否强力单卡', '是否娱乐'
]].copy()
df_full.rename(columns={'标签': '推荐标签'}, inplace=True)

# 推荐子集
df_rec = df_full[df_full['推荐标签'] == '推荐'].copy()

total_rec = len(df_rec)
print(f'\n推荐条目总数: {total_rec}')
print(f'推荐 - 优先级标签分布:')
print(df_rec['优先级标签'].value_counts())

# 英雄维度统计
hero_list = []
for hero in sorted(df_rec['英雄名称'].unique()):
    sub = df_rec[df_rec['英雄名称'] == hero]
    hero_list.append({
        '英雄称号': hero,
        '英雄中文名': title_to_name.get(hero, ''),
        '定位1': title_to_role1.get(hero, ''),
        '定位2': title_to_role2.get(hero, '') if pd.notna(title_to_role2.get(hero)) else '',
        '推荐符文数': len(sub),
        '最佳拍档数': int(sub['是否最佳拍档'].sum()),
        '强力单卡数': int(sub['是否强力单卡'].sum()),
        '娱乐数': int(sub['是否娱乐'].sum()),
        '无标签数': int((sub['优先级标签'] == '无').sum()),
        '标签覆盖率%': round((1 - (sub['优先级标签'] == '无').sum() / len(sub)) * 100, 1),
        '平均表现分': round(sub['表现分'].mean(), 2),
        '平均热度': round(sub['热度'].mean(), 2),
    })
hero_stats = pd.DataFrame(hero_list).sort_values('推荐符文数', ascending=False)

# 符文维度统计
aug_list = []
for aug in sorted(df_rec['符文名称'].unique()):
    sub = df_rec[df_rec['符文名称'] == aug]
    rarity = sub['稀有度'].iloc[0] if len(sub) > 0 else ''
    aug_list.append({
        '符文名称': aug,
        '稀有度': rarity,
        '被推荐英雄数': len(sub),
        '最佳拍档数': int(sub['是否最佳拍档'].sum()),
        '强力单卡数': int(sub['是否强力单卡'].sum()),
        '娱乐数': int(sub['是否娱乐'].sum()),
        '无标签数': int((sub['优先级标签'] == '无').sum()),
        '标签覆盖率%': round((1 - (sub['优先级标签'] == '无').sum() / len(sub)) * 100, 1),
        '平均表现分': round(sub['表现分'].mean(), 2),
        '平均热度': round(sub['热度'].mean(), 2),
    })
aug_stats = pd.DataFrame(aug_list).sort_values('被推荐英雄数', ascending=False)

# 无标签明细
df_no_label = df_rec[df_rec['优先级标签'] == '无'][[
    '英雄名称', '英雄中文名', '定位1', '定位2',
    '符文名称', '符文ID', '稀有度', '评级', '表现分', '热度'
]].sort_values(['英雄名称', '表现分'], ascending=[True, False])

# 总览
summary_rows = [
    ('推荐条目总数（英雄×符文）', total_rec),
    ('涉及英雄数', df_rec['英雄名称'].nunique()),
    ('涉及符文数', df_rec['符文名称'].nunique()),
    ('', ''),
    ('── 标签覆盖（同一条目可多标签） ──', ''),
    ('最佳拍档覆盖条目数', int(df_rec['是否最佳拍档'].sum())),
    ('最佳拍档覆盖占比', f"{df_rec['是否最佳拍档'].sum()/total_rec*100:.1f}%"),
    ('强力单卡覆盖条目数', int(df_rec['是否强力单卡'].sum())),
    ('强力单卡覆盖占比', f"{df_rec['是否强力单卡'].sum()/total_rec*100:.1f}%"),
    ('娱乐覆盖条目数', int(df_rec['是否娱乐'].sum())),
    ('娱乐覆盖占比', f"{df_rec['是否娱乐'].sum()/total_rec*100:.1f}%"),
    ('', ''),
    ('── 有/无标签统计 ──', ''),
    ('有标签条目数（任一标签命中）', total_rec - int((df_rec['优先级标签'] == '无').sum())),
    ('有标签占比', f"{(1 - (df_rec['优先级标签'] == '无').sum()/total_rec)*100:.1f}%"),
    ('无标签条目数', int((df_rec['优先级标签'] == '无').sum())),
    ('无标签占比', f"{(df_rec['优先级标签'] == '无').sum()/total_rec*100:.1f}%"),
    ('', ''),
    ('── 优先级标签分布（互斥，优先级：最佳拍档>娱乐>强力单卡） ──', ''),
    ('  最佳拍档', int((df_rec['优先级标签'] == '最佳拍档').sum())),
    ('  强力单卡', int((df_rec['优先级标签'] == '强力单卡').sum())),
    ('  娱乐', int((df_rec['优先级标签'] == '娱乐').sum())),
    ('  无标签', int((df_rec['优先级标签'] == '无').sum())),
]
df_summary = pd.DataFrame(summary_rows, columns=['指标', '数值'])

print(f'\n英雄统计: {len(hero_stats)} 行')
print(f'符文统计: {len(aug_stats)} 行')
print(f'无标签明细: {len(df_no_label)} 行')

# ============================================================
# 写入Excel
# ============================================================
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='总览', index=False)
    df_rec.to_excel(writer, sheet_name='推荐符文_含标签', index=False)
    hero_stats.to_excel(writer, sheet_name='英雄维度统计', index=False)
    aug_stats.to_excel(writer, sheet_name='符文维度统计', index=False)
    df_no_label.to_excel(writer, sheet_name='无标签明细', index=False)
    df_full.to_excel(writer, sheet_name='完整预选池_含标签', index=False)

# ============================================================
# 美化格式
# ============================================================
wb = load_workbook(output_path)

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
data_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

label_fills = {
    '最佳拍档': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    '强力单卡': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    '娱乐': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
    '无': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
}

for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    
    # 表头格式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 数据行格式（限制处理行数避免太慢）
    max_format_rows = min(ws.max_row + 1, 8000)
    for row in range(2, max_format_rows):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 自动列宽
    for col in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ''))
        for row in range(2, min(ws.max_row + 1, 50)):
            val_len = len(str(ws.cell(row=row, column=col).value or ''))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)
    
    # 自动筛选
    if ws.max_row > 1 and ws.title != '总览':
        ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# 为推荐符文表的"优先级标签"列上色
ws_rec = wb['推荐符文_含标签']
label_col = None
for col in range(1, ws_rec.max_column + 1):
    if ws_rec.cell(row=1, column=col).value == '优先级标签':
        label_col = col
        break

if label_col:
    for row in range(2, min(ws_rec.max_row + 1, 8000)):
        cell = ws_rec.cell(row=row, column=label_col)
        val = str(cell.value or '')
        if val in label_fills:
            cell.fill = label_fills[val]

# 为无标签明细表添加浅黄底色
ws_no = wb['无标签明细']
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
for row in range(2, min(ws_no.max_row + 1, 2000)):
    for col in range(1, ws_no.max_column + 1):
        ws_no.cell(row=row, column=col).fill = yellow_fill

# 总览页签特殊处理
ws_ov = wb['总览']
ws_ov.column_dimensions['A'].width = 50
ws_ov.column_dimensions['B'].width = 18
section_font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')
for row in range(2, ws_ov.max_row + 1):
    val = str(ws_ov.cell(row=row, column=1).value or '')
    if val.startswith('──'):
        ws_ov.cell(row=row, column=1).font = section_font

wb.save(output_path)
print(f'\n✅ 已输出到: {output_path}')
print(f'包含页签: 总览 / 推荐符文_含标签 / 英雄维度统计 / 符文维度统计 / 无标签明细 / 完整预选池_含标签')
